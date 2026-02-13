# prtg_report.py - Fetch PRTG sensor data and generate Excel report in one step
import requests
import pandas as pd
import numpy as np
import json
import sys
import os
from datetime import datetime, timedelta
from io import StringIO
import urllib3

urllib3.disable_warnings()

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("✗ openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

# ============================================================
# LOAD CONFIGURATION FROM config.json
# ============================================================

CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')

try:
    with open(CONFIG_FILE, 'r') as f:
        config = json.load(f)
except FileNotFoundError:
    print(f"✗ Config file not found: {CONFIG_FILE}")
    sys.exit(1)

PRTG_CONFIG = config['prtg']
SENSORS = {int(k): v for k, v in config['sensors'].items()}
DAYS_TO_ANALYZE = config.get('days_to_analyze', 2)

# ============================================================
# EXCEL STYLES
# ============================================================

BG_DARK = '1a1a2e'
BG_CARD = '16213e'
BG_HEADER = '0f3460'
ACCENT = '00d4ff'
GREEN = '00e676'
YELLOW = 'ffab00'
ORANGE = 'ff6d00'
RED = 'ff1744'
WHITE = 'ffffff'
DIM = '8892a0'

HEADER_FONT = Font(name='Calibri', bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=BG_HEADER, end_color=BG_HEADER, fill_type='solid')
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color=ACCENT)
SUBTITLE_FONT = Font(name='Calibri', size=10, color=DIM)
DATA_FONT = Font(name='Calibri', size=11, color=WHITE)
DATA_BOLD = Font(name='Calibri', bold=True, size=11, color=WHITE)
CARD_FILL = PatternFill(start_color=BG_CARD, end_color=BG_CARD, fill_type='solid')
DARK_FILL = PatternFill(start_color=BG_DARK, end_color=BG_DARK, fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color=BG_HEADER),
    right=Side(style='thin', color=BG_HEADER),
    top=Side(style='thin', color=BG_HEADER),
    bottom=Side(style='thin', color=BG_HEADER),
)

def colored_font(color, bold=False, size=11):
    return Font(name='Calibri', bold=bold, size=size, color=color)

def get_stability(std):
    if std < 1.0:
        return "VERY STABLE", GREEN
    elif std < 2.0:
        return "STABLE", '2979ff'
    elif std < 3.0:
        return "VARIABLE", YELLOW
    else:
        return "HIGHLY VARIABLE", RED

# ============================================================
# PRTG DATA FETCHING
# ============================================================

def fetch_sensor_history(sensor_id):
    """Fetch historical data from PRTG"""
    end_date = datetime.now()
    start_date = end_date - timedelta(days=DAYS_TO_ANALYZE)

    params = {
        'id': sensor_id,
        'sdate': start_date.strftime('%Y-%m-%d-00-00-00'),
        'edate': end_date.strftime('%Y-%m-%d-23-59-59'),
        'avg': 300,
        'username': PRTG_CONFIG['username'],
        'password': PRTG_CONFIG['password'],
    }

    print(f"  Fetching data (last {DAYS_TO_ANALYZE} days)...")

    try:
        response = requests.get(
            f"{PRTG_CONFIG['url']}/api/historicdata.csv",
            params=params,
            verify=False,
            timeout=30
        )

        if response.status_code == 200:
            print(f"  ✓ Retrieved ({len(response.text)} bytes)")
            return response.text
        else:
            print(f"  ✗ HTTP {response.status_code}: {response.text[:150]}")
            return None

    except Exception as e:
        print(f"  ✗ Connection error: {e}")
        return None

def parse_csv_data(csv_data):
    """Parse PRTG CSV data into a clean DataFrame"""
    df = pd.read_csv(StringIO(csv_data))

    # Drop summary rows
    df = df[df['Date Time'].str.match(r'^\d', na=False)].copy()

    # Find temperature column
    tempc_cols = [col for col in df.columns if 'tempc' in col.lower() and '(raw)' not in col.lower()]
    temp_cols = [col for col in df.columns if 'temperature' in col.lower() or '(c)' in col.lower()]

    if tempc_cols:
        temp_col = tempc_cols[0]
    elif temp_cols:
        temp_col = temp_cols[0]
    else:
        print("  ✗ No temperature column found!")
        return None

    print(f"  ✓ Column: '{temp_col}'")

    # Parse datetime
    df['DateTime'] = pd.to_datetime(
        df['Date Time'].str.split(' - ').str[0],
        format='%d.%m.%Y %H:%M:%S'
    )

    # Clean temperature data
    df[temp_col] = df[temp_col].astype(str).str.replace(r'[^0-9.\-]', '', regex=True)
    df[temp_col] = pd.to_numeric(df[temp_col], errors='coerce')
    df = df.dropna(subset=[temp_col])

    print(f"  ✓ {len(df)} valid readings")

    df = df[['DateTime', temp_col]].rename(columns={temp_col: 'Temperature'})
    return df

def compute_stats(df):
    """Compute all stats for a sensor DataFrame"""
    temps = df['Temperature']
    days = (df['DateTime'].max() - df['DateTime'].min()).days

    current = temps.iloc[-1]
    avg = temps.mean()
    mn = temps.min()
    mx = temps.max()
    rng = mx - mn
    std = temps.std()
    stability, stab_color = get_stability(std)

    if days < 7:
        ue = avg + 3.0
        uw = avg + 2.0
        lw = avg - 2.0
        le = avg - 3.0
        threshold_note = "Conservative estimates (limited data)"
    else:
        ue = temps.quantile(0.99)
        uw = temps.quantile(0.95)
        lw = temps.quantile(0.05)
        le = temps.quantile(0.01)
        threshold_note = "Based on statistical analysis"

    return {
        'df': df, 'days': days,
        'current': current, 'avg': avg, 'min': mn, 'max': mx,
        'range': rng, 'std': std,
        'stability': stability, 'stab_color': stab_color,
        'ue': ue, 'uw': uw, 'lw': lw, 'le': le,
        'threshold_note': threshold_note,
        'p01': temps.quantile(0.01), 'p05': temps.quantile(0.05),
        'p25': temps.quantile(0.25), 'p50': temps.quantile(0.50),
        'p75': temps.quantile(0.75), 'p95': temps.quantile(0.95),
        'p99': temps.quantile(0.99),
    }

# ============================================================
# EXCEL REPORT GENERATION
# ============================================================

def write_summary_sheet(ws, sensor_results):
    """Sheet 1: Summary table of all sensors"""
    # Dark background
    for row in range(1, 50):
        for col in range(1, 20):
            ws.cell(row=row, column=col).fill = DARK_FILL

    # Column widths
    for i, w in enumerate([22, 12, 10, 10, 10, 10, 10, 14, 12, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells('A1:L1')
    c = ws['A1']
    c.value = 'PRTG Temperature Sensor Report'
    c.font = Font(name='Calibri', bold=True, size=18, color=ACCENT)
    c.fill = DARK_FILL
    c.alignment = LEFT

    ws.merge_cells('A2:L2')
    c = ws['A2']
    c.value = datetime.now().strftime('%B %d, %Y  %H:%M')
    c.font = SUBTITLE_FONT
    c.fill = DARK_FILL

    # Headers
    headers = ['Sensor', 'Current', 'Avg', 'Min', 'Max', 'Range',
               'Std Dev', 'Stability', 'Upper Err', 'Upper Warn',
               'Lower Warn', 'Lower Err']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Data rows
    row_num = 5
    for sid, (name, s) in sensor_results.items():
        # Current temp color
        if s['current'] >= s['ue'] or s['current'] <= s['le']:
            cur_color = RED
        elif s['current'] >= s['uw'] or s['current'] <= s['lw']:
            cur_color = ORANGE
        else:
            cur_color = GREEN

        row_fill = CARD_FILL if (row_num - 5) % 2 == 0 else DARK_FILL

        values = [
            (f'{name} ({sid})', DATA_BOLD, LEFT),
            (round(s['current'], 1), colored_font(cur_color, bold=True), CENTER),
            (round(s['avg'], 1), DATA_FONT, CENTER),
            (round(s['min'], 1), DATA_FONT, CENTER),
            (round(s['max'], 1), DATA_FONT, CENTER),
            (round(s['range'], 1), DATA_FONT, CENTER),
            (round(s['std'], 2), DATA_FONT, CENTER),
            (s['stability'], colored_font(s['stab_color'], bold=True), CENTER),
            (round(s['ue'], 1), colored_font(RED), CENTER),
            (round(s['uw'], 1), colored_font(ORANGE), CENTER),
            (round(s['lw'], 1), colored_font(ORANGE), CENTER),
            (round(s['le'], 1), colored_font(RED), CENTER),
        ]

        for col, (val, font, align) in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = font
            cell.fill = row_fill
            cell.alignment = align
            cell.border = THIN_BORDER

        row_num += 1

    # Footer
    row_num += 1
    ws.merge_cells(f'A{row_num}:L{row_num}')
    c = ws.cell(row=row_num, column=1)
    c.value = f'{len(sensor_results)} sensor(s)  |  Prepared by: Data Center Operations'
    c.font = Font(name='Calibri', size=9, color=DIM)
    c.fill = DARK_FILL

def write_detailed_sheet(ws, sid, name, s):
    """One sheet per sensor: stats, percentiles, hourly averages"""
    ws.sheet_properties.tabColor = '2979ff'

    # Dark background
    for row in range(1, 80):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = DARK_FILL

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    for i in range(3, 8):
        ws.column_dimensions[get_column_letter(i)].width = 14

    df = s['df']

    # Sensor title
    row_num = 1
    ws.merge_cells(f'A{row_num}:D{row_num}')
    c = ws.cell(row=row_num, column=1)
    c.value = f"{name} (Sensor {sid})"
    c.font = Font(name='Calibri', bold=True, size=14, color=ACCENT)
    c.fill = DARK_FILL
    row_num += 1

    c = ws.cell(row=row_num, column=1)
    c.value = f"{s['days']} days  |  {len(df)} readings"
    c.font = SUBTITLE_FONT
    c.fill = DARK_FILL
    row_num += 2

    # Stats
    stat_rows = [
        ('Current Temperature', f"{s['current']:.2f}°C"),
        ('Average', f"{s['avg']:.2f}°C"),
        ('Minimum', f"{s['min']:.2f}°C"),
        ('Maximum', f"{s['max']:.2f}°C"),
        ('Range', f"{s['range']:.2f}°C"),
        ('Standard Deviation', f"{s['std']:.2f}°C"),
        ('Stability', s['stability']),
        ('', ''),
        ('PERCENTILES', ''),
        ('1st percentile', f"{s['p01']:.2f}°C"),
        ('5th percentile', f"{s['p05']:.2f}°C"),
        ('25th percentile', f"{s['p25']:.2f}°C"),
        ('50th percentile (median)', f"{s['p50']:.2f}°C"),
        ('75th percentile', f"{s['p75']:.2f}°C"),
        ('95th percentile', f"{s['p95']:.2f}°C"),
        ('99th percentile', f"{s['p99']:.2f}°C"),
        ('', ''),
        ('THRESHOLDS', ''),
        ('Upper Error', f"{s['ue']:.1f}°C"),
        ('Upper Warning', f"{s['uw']:.1f}°C"),
        ('Lower Warning', f"{s['lw']:.1f}°C"),
        ('Lower Error', f"{s['le']:.1f}°C"),
        ('', ''),
        ('Note', s['threshold_note']),
    ]

    for label, val in stat_rows:
        is_section = label in ('PERCENTILES', 'THRESHOLDS')
        cl = ws.cell(row=row_num, column=1, value=label)
        cv = ws.cell(row=row_num, column=2, value=val)

        if is_section:
            cl.font = colored_font(ACCENT, bold=True)
        elif label == 'Stability':
            cl.font = DATA_FONT
            cv.font = colored_font(s['stab_color'], bold=True)
        elif 'Error' in label:
            cl.font = DATA_FONT
            cv.font = colored_font(RED)
        elif 'Warning' in label:
            cl.font = DATA_FONT
            cv.font = colored_font(ORANGE)
        else:
            cl.font = DATA_FONT
            cv.font = DATA_FONT

        cl.fill = CARD_FILL if not is_section else DARK_FILL
        cv.fill = CARD_FILL if not is_section else DARK_FILL
        cl.alignment = LEFT
        cv.alignment = CENTER
        cl.border = THIN_BORDER
        cv.border = THIN_BORDER
        row_num += 1

    # Hourly averages
    row_num += 1
    c = ws.cell(row=row_num, column=1, value='HOURLY AVERAGES')
    c.font = colored_font(ACCENT, bold=True)
    c.fill = DARK_FILL
    row_num += 1

    df_h = df.copy()
    df_h['Hour'] = df_h['DateTime'].dt.hour
    hourly = df_h.groupby('Hour')['Temperature'].agg(['mean', 'min', 'max'])

    for h_col, h_label in enumerate(['Hour', 'Avg', 'Min', 'Max'], 1):
        cell = ws.cell(row=row_num, column=h_col, value=h_label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    row_num += 1

    for hour in range(24):
        if hour in hourly.index:
            r_fill = CARD_FILL if hour % 2 == 0 else DARK_FILL
            for h_col, val in enumerate([f'{hour:02d}:00',
                                          f"{hourly.loc[hour, 'mean']:.1f}",
                                          f"{hourly.loc[hour, 'min']:.1f}",
                                          f"{hourly.loc[hour, 'max']:.1f}"], 1):
                cell = ws.cell(row=row_num, column=h_col, value=val)
                cell.font = DATA_FONT
                cell.fill = r_fill
                cell.alignment = CENTER
                cell.border = THIN_BORDER
            row_num += 1

def write_raw_sheets(writer, sensor_results):
    """One sheet per sensor with raw DateTime + Temperature data"""
    for sid, (name, s) in sensor_results.items():
        sheet_name = f"{name[:20]} ({sid})"
        df = s['df'].copy()
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

        ws = writer.sheets[sheet_name]
        ws.sheet_properties.tabColor = BG_HEADER

        # Title
        ws.merge_cells('A1:B1')
        c = ws['A1']
        c.value = f"{name} - Raw Data"
        c.font = TITLE_FONT
        c.fill = DARK_FILL

        # Style header
        for col in range(1, 3):
            cell = ws.cell(row=2, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 14

        # Style data rows
        for row in range(3, len(df) + 3):
            r_fill = CARD_FILL if row % 2 == 0 else DARK_FILL
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.font = DATA_FONT
                cell.fill = r_fill
                cell.alignment = CENTER
                cell.border = THIN_BORDER

# ============================================================
# MAIN
# ============================================================

def main():
    print("="*70)
    print("PRTG SENSOR REPORT")
    print("="*70)

    # Step 1: Fetch and parse all sensor data
    sensor_results = {}  # {sid: (name, stats_dict)}

    for sid, name in SENSORS.items():
        print(f"\n[{name}] (Sensor {sid})")

        csv_data = fetch_sensor_history(sid)
        if not csv_data:
            continue

        df = parse_csv_data(csv_data)
        if df is None or len(df) == 0:
            print(f"  ✗ No valid data. Skipping.")
            continue

        stats = compute_stats(df)
        sensor_results[sid] = (name, stats)

        # Print quick summary to console
        print(f"  Current: {stats['current']:.1f}°C  |  "
              f"Avg: {stats['avg']:.1f}°C  |  "
              f"Range: {stats['min']:.1f}-{stats['max']:.1f}°C  |  "
              f"{stats['stability']}")

    if not sensor_results:
        print("\n✗ No sensor data retrieved.")
        return

    # Step 2: Generate Excel report
    print(f"\n{'─'*70}")
    print("Generating Excel report...")

    output_file = f"sensor_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        ws_sum = writer.book.create_sheet('Summary', 0)
        ws_sum.sheet_properties.tabColor = ACCENT
        write_summary_sheet(ws_sum, sensor_results)

        for sid, (name, s) in sensor_results.items():
            detail_name = f"Stats - {name[:22]}"
            ws_det = writer.book.create_sheet(detail_name)
            write_detailed_sheet(ws_det, sid, name, s)

        write_raw_sheets(writer, sensor_results)

        if 'Sheet' in writer.book.sheetnames:
            del writer.book['Sheet']

    print(f"\n✓ Report saved: {output_file}")
    print(f"  {len(sensor_results)} sensor(s) included.")
    print(f"{'='*70}")

if __name__ == "__main__":
    main()
